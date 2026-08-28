"""Tests for `arcelle_sidecar.docs.dispatch` (port of
`src-tauri/src/extraction.rs` lines 29-58, 284-371, 383-396, 398-403 -- the
`extract_text` top-level dispatcher).

This module is pure WIRING: every reader `extract_text` calls already has
its own full test suite elsewhere in `tests/`. These tests therefore focus
on what is unique to the dispatcher itself: the extension table, the
`extension_of` edge cases (verified against real `rustc` output -- see
`extension_of`'s docstring), the fast text-extension path vs. the
extension-less sniff path (and the exact byte-content distinction between
them), every dispatch branch being reached with a real minimal fixture, the
lossy-UTF-8-only vs. full-`decode_text_bytes` asymmetry between html/htm and
rtf/eml/srt/vtt/svg, and the two containment layers (`_contain_parser_panic`,
and the final `normalize_whitespace` + blank filter).
"""

from __future__ import annotations

import io
import json
import zipfile

import openpyxl
import pymupdf
import pytest

from arcelle_sidecar.docs import dispatch
from test_docs_article import NEWS
from test_docs_legacy import _container, _ole_with, _text_chars, _utf16le
from arcelle_sidecar.docs import legacy as legacy_mod

# ------------------------------------------------------------- zip fixtures


def _fake_office_zip(entry: str, xml: str) -> bytes:
    """A single-entry, deflate-compressed Office-style zip -- same pattern
    used throughout `tests/test_docs_*.py`.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(entry, xml)
    return buf.getvalue()


def _build_zip(entries: dict[str, bytes]) -> bytes:
    """An in-memory zip with one entry per (name, bytes) pair."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as writer:
        for name, content in entries.items():
            writer.writestr(name, content)
    return buf.getvalue()


def _pdf_bytes(text: str) -> bytes:
    """A minimal real single-page PDF containing `text`."""
    doc = pymupdf.open()
    try:
        page = doc.new_page(width=400, height=200)
        page.insert_text((20, 100), text, fontsize=24)
        return doc.tobytes()
    finally:
        doc.close()


# =============================================================================
# extension table / text_extensions() / is_text_extension()
# =============================================================================


def test_text_extensions_is_the_exact_rust_list_in_order() -> None:
    # Transcribed directly from `src-tauri/src/extraction.rs` lines 29-35.
    # NOTE: the porting brief's prose called this a "45-entry list", but the
    # Rust source itself (read in full, and re-counted by hand) has 59
    # entries -- this is the exact, verbatim, in-order copy of what the Rust
    # source actually contains.
    expected = (
        "txt", "md", "markdown", "json", "jsonl", "ndjson", "csv", "tsv", "log", "xml", "yml",
        "yaml", "toml", "ini", "rs", "py", "js", "jsx", "ts", "tsx", "java", "c", "h", "cpp",
        "hpp", "cs", "go", "rb", "php", "swift", "kt", "sh", "zsh", "bash", "sql", "r", "m",
        "scala", "lua", "pl", "css", "scss", "less", "vue", "svelte", "tex", "org", "rst",
        "diff", "patch", "dockerfile", "graphql", "gql", "proto", "properties", "env",
        "gitignore", "cfg", "conf",
    )
    assert dispatch.TEXT_EXTENSIONS == expected
    assert len(dispatch.TEXT_EXTENSIONS) == 59
    assert len(dispatch.TEXT_EXTENSIONS) == len(set(dispatch.TEXT_EXTENSIONS)), "duplicate entry"
    assert dispatch.text_extensions() == dispatch.TEXT_EXTENSIONS


@pytest.mark.parametrize("ext", ["py", "json", "dockerfile", "env", "txt", "rst"])
def test_is_text_extension_true_for_table_entries(ext: str) -> None:
    assert dispatch.is_text_extension(ext) is True


@pytest.mark.parametrize("ext", ["exe", "pdf", "docx", ""])
def test_is_text_extension_false_for_non_table_entries(ext: str) -> None:
    assert dispatch.is_text_extension(ext) is False


def test_is_image() -> None:
    assert dispatch.is_image("image/png") is True
    assert dispatch.is_image("image/jpeg") is True
    assert dispatch.is_image("application/pdf") is False
    assert dispatch.is_image("text/plain") is False
    assert dispatch.is_image("") is False


# =============================================================================
# extension_of -- verified against real rustc output on Path::extension()
# =============================================================================


@pytest.mark.parametrize(
    ("name", "expected"),
    [
        (".bashrc", ""),
        ("..foo", "foo"),
        ("a.b.c", "c"),
        ("noext", ""),
        ("trailing.", ""),
        ("web.dockerfile", "dockerfile"),
        ("app.env", "env"),
        (".gitignore", ""),
        ("Dockerfile", ""),
        ("dir/sub/report.PDF", "pdf"),
        ("/a/b/c.PDF", "pdf"),
        ("dir/sub.TXT", "txt"),
        ("dir/.bashrc", ""),
        ("...", ""),
        ("a.", ""),
        ("..", ""),
        (".", ""),
        ("", ""),
        ("a..b", "b"),
        ("..a..b", "b"),
        (".config.bak", "bak"),
        ("  spaced.name.EXT", "ext"),
        ("FOO.TXT", "txt"),
    ],
)
def test_extension_of_matches_rust_path_extension_semantics(name: str, expected: str) -> None:
    assert dispatch.extension_of(name) == expected


# =============================================================================
# fast path (TEXT_EXTENSIONS) round-trips through decode_text_bytes
# =============================================================================


@pytest.mark.parametrize(
    "name",
    ["module.py", "data.json", "web.dockerfile", "app.env", "readme.txt", "notes.markdown"],
)
def test_text_extension_fast_path_round_trips_spot_checks(name: str) -> None:
    data = "line one\nline two\n".encode()
    result = dispatch.extract_text(name, data)
    assert result == "line one\nline two\n"


def test_text_extension_fast_path_bypasses_normalize_whitespace() -> None:
    # The fast path returns `decode_text_bytes(data)` directly -- it must
    # NOT be run through `normalize_whitespace`, unlike everything that goes
    # through `_contain_parser_panic`. Irregular internal whitespace that
    # `normalize_whitespace` would collapse must survive verbatim.
    data = b"a    b\tc"
    assert dispatch.extract_text("weird.py", data) == "a    b\tc"


# =============================================================================
# the dotted-extension vs. bare-stem distinction (the historical gotcha)
# =============================================================================


def test_dotted_extension_hits_fast_path_bare_stem_hits_sniff_path() -> None:
    # Same bytes, two different filenames. A NUL byte inside otherwise
    # valid-UTF-8-shaped bytes is exactly the signal `_sniff_text_bytes`
    # refuses (a strong hint of a binary file) -- but the TEXT_EXTENSIONS
    # fast path does not gate on byte content at all, so a DOTTED name with
    # a real extension reads it fine while the BARE stem (ext == "") does
    # not.
    data = b"KEY=value\x00TRAILING_JUNK"

    assert dispatch.extract_text("app.env", data) is not None
    assert dispatch.extract_text(".env", data) is None

    assert dispatch.extract_text("web.dockerfile", data) is not None
    assert dispatch.extract_text("Dockerfile", data) is None

    assert dispatch.extract_text("thing.gitignore", data) is not None
    assert dispatch.extract_text(".gitignore", data) is None


def test_bare_dockerfile_stem_uses_the_sniff_path_when_content_is_clean() -> None:
    # A bare "Dockerfile" (no dot at all) has ext == "" -- NOT in
    # TEXT_EXTENSIONS (which only ever holds the lowercase word
    # "dockerfile" as an EXTENSION, matched by `web.dockerfile`-shaped
    # names). Clean sniffable content still comes back as text, via the
    # no-extension sniff path rather than the fast path.
    data = b"FROM python:3.12\nRUN pip install stuff\n"
    assert dispatch.extension_of("Dockerfile") == ""
    text = dispatch.extract_text("Dockerfile", data)
    assert text is not None
    assert "FROM python:3.12" in text


def test_bare_dotenv_stem_uses_the_sniff_path() -> None:
    data = b"API_KEY=xyz\n"
    assert dispatch.extension_of(".env") == ""
    text = dispatch.extract_text(".env", data)
    assert text is not None
    assert "API_KEY=xyz" in text


# =============================================================================
# the extension-less sniff path
# =============================================================================


def test_no_extension_plain_sniffable_text_is_read_verbatim() -> None:
    data = b"Some plain README prose.\nSecond line.\n"
    result = dispatch.extract_text("README", data)
    assert result == "Some plain README prose.\nSecond line.\n"


def test_no_extension_with_embedded_nul_is_not_sniffed() -> None:
    # Valid UTF-8 shape, but a literal NUL inside it -- the strong binary
    # signal `_sniff_text_bytes` refuses even though it decodes cleanly.
    data = b"looks\x00like\x00text"
    assert dispatch.extract_text("payload", data) is None


def test_no_extension_bytes_that_only_decode_via_detection_are_not_sniffed() -> None:
    # windows-1252-ish high-bit bytes: not a valid UTF-8 sequence and not a
    # recognised BOM -- can only be read via `EncodingSource.DETECTED`,
    # which the sniff path must reject outright regardless of the guess.
    data = bytes((0xE9, 0xE8, 0xE0, 0xF9)) * 20
    detail = dispatch.decode_text_detail(data)
    assert detail.source == dispatch.EncodingSource.DETECTED, "fixture must actually detect"
    assert dispatch.extract_text("mystery", data) is None


def test_no_extension_bom_full_of_nuls_is_still_sniffed() -> None:
    # UTF-16LE BOM + ASCII text: every other byte is a NUL by construction,
    # but a BOM is a FACT about the encoding, not a guess -- the NUL check
    # only applies to the UTF8 source, never to BOM.
    data = b"\xff\xfe" + "Hello there".encode("utf-16-le")
    result = dispatch.extract_text("mystery", data)
    assert result == "Hello there"


def test_no_extension_lossy_bom_decode_is_not_sniffed() -> None:
    # UTF-16LE BOM followed by an odd trailing byte: a truncated code unit,
    # which forces a lossy (errors="replace") decode. A BOM is accepted only
    # when the decode is NOT lossy.
    data = b"\xff\xfe" + "Hi".encode("utf-16-le") + b"\x41"
    assert dispatch.extract_text("mystery", data) is None


def test_no_extension_blank_after_strip_is_not_sniffed() -> None:
    assert dispatch.extract_text("mystery", b"   \n\t  \n  ") is None


def test_no_extension_falls_through_to_none_when_sniff_fails() -> None:
    # ext == "" also matches no arm in the extension dispatch, so a failed
    # sniff must end in None, not an exception.
    data = bytes((0xE9, 0xE8))
    assert dispatch.extract_text("mystery", data) is None


# =============================================================================
# every dispatch branch is reached, with a real minimal fixture
# =============================================================================


def test_pdf_branch() -> None:
    data = _pdf_bytes("Hello from a real PDF page.")
    result = dispatch.extract_text("report.pdf", data)
    assert result is not None
    assert "Hello from a real PDF page." in result


def test_uppercase_extension_dispatches_through_the_lowercased_branch() -> None:
    # Adversarial case: `extension_of` lowercases via Rust's own
    # `to_lowercase()` on `extension_of`, and `extract_text` computes `ext`
    # exactly once up front -- every later comparison (the TEXT_EXTENSIONS
    # containment check AND every arm of the extension match) reads that
    # same lowercased `ext`, so the dispatch table itself must never see a
    # raw-cased extension. If a future edit dispatched on the ORIGINAL
    # (un-lowercased) name instead, "Report.PDF" would silently fall through
    # to the `_ => None` arm -- indistinguishable from a genuinely unreadable
    # file, which is exactly the failure mode this test exists to catch.
    data = _pdf_bytes("Hello uppercase PDF.")
    result = dispatch.extract_text("Report.PDF", data)
    assert result is not None
    assert "Hello uppercase PDF." in result

    # Same adversarial check on the OTHER dispatch path -- the TEXT_EXTENSIONS
    # fast path -- which does its own independent `ext in TEXT_EXTENSIONS`
    # containment check on the same lowercased `ext`.
    text_result = dispatch.extract_text("NOTES.TXT", b"hello uppercase name")
    assert text_result == "hello uppercase name"


def test_docx_branch() -> None:
    xml = "<w:document><w:body><w:p><w:r><w:t>Hello docx body.</w:t></w:r></w:p></w:body></w:document>"
    data = _fake_office_zip("word/document.xml", xml)
    result = dispatch.extract_text("letter.docx", data)
    assert result is not None
    assert "Hello docx body." in result


def test_xlsx_branch() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Quarterly total"
    buf = io.BytesIO()
    wb.save(buf)
    result = dispatch.extract_text("budget.xlsx", buf.getvalue())
    assert result is not None
    assert "Quarterly total" in result


def test_pptx_branch() -> None:
    xml = '<p:sld><a:t>Hello pptx headline.</a:t><a:p></a:p></p:sld>'
    data = _fake_office_zip("ppt/slides/slide1.xml", xml)
    result = dispatch.extract_text("deck.pptx", data)
    assert result is not None
    assert "Hello pptx headline." in result


def test_html_branch_returns_the_article_when_one_is_found() -> None:
    # NEWS is the exact fixture `test_docs_article.py` already proved scores
    # as a real article under the readability engine -- reused rather than
    # hand-rolling a page and hoping it clears the scoring threshold.
    text = dispatch.extract_text("saved.html", NEWS.encode("utf-8"))
    assert text is not None
    assert "After eleven years of absence" in text
    assert "Subscribe now" not in text, "chrome leaked into the dispatched article text"


def test_htm_alias_falls_back_to_strip_html_when_there_is_no_article() -> None:
    shell = (
        '<html><head><title>Results</title></head><body>'
        '<nav>a b c</nav><div id="app">no readable article here</div></body></html>'
    )
    text = dispatch.extract_text("page.htm", shell.encode("utf-8"))
    assert text is not None
    assert "no readable article here" in text


def test_epub_branch() -> None:
    data = _build_zip(
        {"OEBPS/aaa.xhtml": b"<html><body><p>Hello epub chapter.</p></body></html>"}
    )
    result = dispatch.extract_text("book.epub", data)
    assert result is not None
    assert "Hello epub chapter." in result


def test_rtf_branch() -> None:
    rtf_doc = r"{\rtf1\ansi\ansicpg1252\deff0 Hello rtf world\par}"
    result = dispatch.extract_text("note.rtf", rtf_doc.encode())
    assert result is not None
    assert "Hello rtf world" in result


def test_doc_branch() -> None:
    data = _ole_with("WordDocument", _utf16le("Hello legacy doc body.\r"))
    result = dispatch.extract_text("old.doc", data)
    assert result is not None
    assert "Hello legacy doc body." in result


def test_ppt_branch() -> None:
    slide = _container(legacy_mod.RT_SLIDE, _text_chars("Hello legacy slide text.\r"))
    data = _ole_with("PowerPoint Document", slide)
    result = dispatch.extract_text("old.ppt", data)
    assert result is not None
    assert "Hello legacy slide text." in result


def test_xls_branch() -> None:
    xlwt = pytest.importorskip("xlwt")
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "Hello xls cell")
    buf = io.BytesIO()
    wb.save(buf)
    result = dispatch.extract_text("book.xls", buf.getvalue())
    assert result is not None
    assert "Hello xls cell" in result


def test_ods_branch() -> None:
    odf_table = pytest.importorskip("odf.table")
    odf_text = pytest.importorskip("odf.text")
    from odf.opendocument import OpenDocumentSpreadsheet

    Table, TableRow, TableCell = odf_table.Table, odf_table.TableRow, odf_table.TableCell
    P = odf_text.P

    doc = OpenDocumentSpreadsheet()
    table = Table(name="Sheet1")
    row = TableRow()
    cell = TableCell(valuetype="string")
    cell.addElement(P(text="Hello ods cell"))
    row.addElement(cell)
    table.addElement(row)
    doc.spreadsheet.addElement(table)
    buf = io.BytesIO()
    doc.write(buf)

    result = dispatch.extract_text("book.ods", buf.getvalue())
    assert result is not None
    assert "Hello ods cell" in result


@pytest.mark.parametrize("name", ["deck.key", "sheet.numbers", "doc.pages"])
def test_iwork_branches(name: str) -> None:
    pdf = _pdf_bytes("Hello iwork preview.")
    data = _build_zip({"QuickLook/Preview.pdf": pdf})
    result = dispatch.extract_text(name, data)
    assert result is not None
    assert "Hello iwork preview." in result


def test_ipynb_branch() -> None:
    nb = {"cells": [{"cell_type": "markdown", "source": ["# Title\n", "Hello notebook body.\n"]}]}
    data = json.dumps(nb).encode()
    result = dispatch.extract_text("analysis.ipynb", data)
    assert result is not None
    assert "Hello notebook body." in result


def test_eml_branch() -> None:
    eml = (
        "From: a@example.com\r\nTo: b@example.com\r\nSubject: Hello eml subject\r\n\r\n"
        "Hello eml body.\r\n"
    )
    result = dispatch.extract_text("msg.eml", eml.encode())
    assert result is not None
    assert "Hello eml subject" in result
    assert "Hello eml body." in result


@pytest.mark.parametrize("name", ["captions.srt", "captions.vtt"])
def test_srt_and_vtt_branches(name: str) -> None:
    if name.endswith(".srt"):
        raw = "1\n00:00:01,000 --> 00:00:04,000\nHello subtitle text.\n\n"
    else:
        raw = "WEBVTT\n\n00:01.000 --> 00:02.000\nHello subtitle text.\n"
    result = dispatch.extract_text(name, raw.encode())
    assert result is not None
    assert "Hello subtitle text." in result


def test_svg_branch() -> None:
    svg = '<svg><title>Hello svg title</title><text x="1">Hello svg body</text></svg>'
    result = dispatch.extract_text("chart.svg", svg.encode())
    assert result is not None
    assert "Hello svg title" in result
    assert "Hello svg body" in result


def test_zip_branch() -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("reports/q1.txt", b"ignored")
        zf.writestr("reports/q2.txt", b"ignored")
    result = dispatch.extract_text("archive.zip", buf.getvalue())
    assert result is not None
    assert "reports/q1.txt" in result
    assert "reports/q2.txt" in result


def test_sketch_always_returns_none() -> None:
    # DELIBERATE, DOCUMENTED gap: sketch extraction depends on the in-house
    # drawing-format module (`commands::sketchdoc`), which has not been
    # ported to this Python sidecar yet.
    assert dispatch.extract_text("drawing.sketch", b'{"anything": true}') is None
    assert dispatch.extract_text("drawing.sketch", b"") is None


def test_unknown_extension_returns_none() -> None:
    assert dispatch.extract_text("mystery.xyz123", b"some content here") is None
    assert dispatch.extract_text("archive.rar", b"some content here") is None


# =============================================================================
# panic containment
# =============================================================================


def test_reader_raising_an_unexpected_exception_is_contained(monkeypatch: pytest.MonkeyPatch) -> None:
    def _boom(data: bytes) -> str | None:
        raise RuntimeError("simulated parser crash")

    monkeypatch.setattr(dispatch, "extract_pdf", _boom)
    assert dispatch.extract_text("crashy.pdf", b"whatever bytes") is None


def test_contain_parser_panic_returns_the_value_on_success() -> None:
    assert dispatch._contain_parser_panic(lambda: "ok") == "ok"
    assert dispatch._contain_parser_panic(lambda: None) is None


def test_contain_parser_panic_contains_any_exception() -> None:
    def _raise() -> str | None:
        raise ValueError("boom")

    assert dispatch._contain_parser_panic(_raise) is None


def test_contain_parser_panic_does_not_catch_keyboard_interrupt() -> None:
    def _interrupt() -> str | None:
        raise KeyboardInterrupt

    with pytest.raises(KeyboardInterrupt):
        dispatch._contain_parser_panic(_interrupt)


# =============================================================================
# the final normalize_whitespace + blank filter, on the contain_parser_panic path
# =============================================================================


def test_final_normalize_whitespace_pass_actually_runs(monkeypatch: pytest.MonkeyPatch) -> None:
    raw = "  Hello   there  \n\n\n\n  General   Kenobi  \n"
    monkeypatch.setattr(dispatch, "extract_zip_listing", lambda data: raw)

    result = dispatch.extract_text("anything.zip", b"ignored")

    assert result == dispatch.normalize_whitespace(raw)
    # And that this is a REAL normalization, not a no-op passthrough.
    assert result != raw
    assert "Hello there" in result
    assert "General Kenobi" in result


def test_extractor_returning_only_whitespace_ends_as_none(monkeypatch: pytest.MonkeyPatch) -> None:
    # The extractor itself returned "something" (a non-None string) -- but
    # it is entirely whitespace, so the final blank filter must still turn
    # it into None.
    monkeypatch.setattr(dispatch, "extract_zip_listing", lambda data: "   \n\n\t  \n   ")
    assert dispatch.extract_text("anything.zip", b"ignored") is None


def test_extractor_returning_none_stays_none(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(dispatch, "extract_zip_listing", lambda data: None)
    assert dispatch.extract_text("anything.zip", b"ignored") is None
