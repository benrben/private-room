"""Tests for `arcelle_sidecar.docs.xlsx` (port of `extract_xlsx` in
`src-tauri/src/extraction/xlsx.rs`).

Mirrors the Rust `#[cfg(test)]` cases: `extracts_xlsx_numeric_and_string_cells`,
`columns_stay_separated_and_blank_cells_keep_their_place`,
`a_sparse_sheet_costs_its_cells_not_its_bounding_rectangle`,
`refuses_workbook_declaring_more_than_cap`,
`refuses_workbook_inflating_more_than_cap`, and
`refuses_workbook_with_lying_declared_sizes` -- plus a handful of
regression tests for divergences this port had to reconcile against the
Rust source (Python's `str(float)` trailing-`.0`, and confirming the
zip-bomb guards actually gate `extract_xlsx` itself, not just the raw
helpers they are built on).

The three decompression-bomb tests mirror the Rust source's OWN structure:
those Rust tests call `zip_declared_size_within`/`zip_inflated_size_within`
directly (not through `extract_xlsx`), because that is the only way to
distinguish "the guard passed" from "the guard failed" without handing
openpyxl an actual multi-hundred-MB workbook -- the guards themselves are
the shared `arcelle_sidecar.docs.xml_utils` functions this module imports
rather than redefines, already covered end-to-end (including this exact
byte-patching attack) by `tests/test_docs_xml_utils.py`. This file re-runs
them here for `xlsx.py`'s own discoverability, and adds one composition
test that patches the real module constant to confirm `extract_xlsx`
itself is wired to refuse, not just the library functions it calls.
"""

from __future__ import annotations

import io
import struct
import time
import zipfile
from types import SimpleNamespace

import openpyxl
import pytest

from arcelle_sidecar.docs import xlsx
from arcelle_sidecar.docs import xml_utils as xu


def _fake_office_zip(entry: str, xml: str) -> bytes:
    """Mirrors the Rust test helper `fake_office_zip`: a single-entry zip
    archive, deflate-compressed like a real Office part.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(entry, xml)
    return buf.getvalue()


def _patch_declared_size(data: bytes, new_size: int) -> bytes:
    """Mirrors the Rust test's exact byte-patching technique
    (`refuses_workbook_with_lying_declared_sizes`): overwrite the declared
    uncompressed-size field in BOTH the local file header (4 bytes at
    offset+22 past `PK\\x03\\x04`) and the central directory record (4
    bytes at offset+24 past `PK\\x01\\x02`) with a little-endian u32, so the
    zip's own metadata lies about how big the entry really inflates.
    """
    out = bytearray(data)
    local = out.index(b"PK\x03\x04")
    out[local + 22 : local + 26] = struct.pack("<I", new_size)
    central = out.index(b"PK\x01\x02")
    out[central + 24 : central + 28] = struct.pack("<I", new_size)
    return bytes(out)


def _workbook_bytes(build) -> bytes:
    """Build a workbook via `build(worksheet)`, save it, and return its
    bytes -- the real-openpyxl-written-xlsx equivalent of the Rust tests'
    `umya_spreadsheet::new_file()` + `write_writer(...)` fixtures.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    build(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------- extract_xlsx


def test_extracts_xlsx_numeric_and_string_cells() -> None:
    # Regression: numeric cells live inline in the worksheet XML, not in
    # sharedStrings.xml. An all-numeric sheet must still extract -- this
    # was the ORIGINAL bug this module exists to fix.
    def build(ws):
        ws["A1"] = "Marketing"
        ws["B1"] = 12000
        ws["A2"] = "Engineering"
        ws["B2"] = 45000

    text = xlsx.extract_xlsx(_workbook_bytes(build))
    assert text is not None
    assert "Marketing" in text, f"got: {text}"
    assert "12000" in text, f"numeric cell missing: {text}"
    assert "45000" in text, f"numeric cell missing: {text}"


def test_columns_stay_separated_and_blank_cells_keep_their_place() -> None:
    # Regression: cells joined with a tab would vanish under
    # `normalize_whitespace` (which collapses tabs), flattening the grid
    # into a run of words with empty cells gone entirely
    # ("Name | (blank) | Total" -> "Name Total") -- exactly how a column
    # gets misread.
    #
    # NOTE on the exact spacing: `extract_xlsx` alone (this module's whole
    # scope) joins ["Name", "", "Total"] with " | ", which -- by ordinary
    # join semantics, verified directly below -- puts TWO spaces between
    # the pipes ("Name |  | Total"), not one. The Rust test asserts a
    # SINGLE space ("Name | | Total") because its assertion runs on
    # `extract_text(...)`'s result, the outer dispatcher that (per
    # `xml_utils.normalize_whitespace`'s own docstring: "this runs over
    # EVERY extractor's output") collapses that double space down to one
    # -- a step that belongs to the dispatcher this port was not asked to
    # build, not to `extract_xlsx` itself. Both facts are checked here so
    # the distinction is on the record rather than silently assumed.
    def build(ws):
        ws["A1"] = "Name"
        ws["C1"] = "Total"  # B1 deliberately blank

    text = xlsx.extract_xlsx(_workbook_bytes(build))
    assert text is not None
    assert "Name |  | Total" in text, f"blank column collapsed: {text}"
    # The literal substring the Rust test (and this port's brief) names,
    # reproduced after the SAME `normalize_whitespace` pass every
    # extractor's output runs through one layer up.
    assert "Name | | Total" in xu.normalize_whitespace(text)


def test_a_sparse_sheet_costs_its_cells_not_its_bounding_rectangle() -> None:
    # Regression: extraction must walk only POPULATED cells, not every
    # coordinate in the rows x columns rectangle. A workbook holding two
    # values whose highest cell sits far out spans tens of millions of
    # coordinates -- a bounding-rectangle walk cost seconds here (measured
    # ~1.0s for the naive `iter_rows(values_only=True)` translation, see
    # xlsx.py's module docstring); this walk (via the internal cell-only
    # parser) measures ~0.1ms for the identical scenario.
    def build(ws):
        ws["A1"] = "Alpha"
        ws.cell(row=40_000, column=1000, value="Omega")

    data = _workbook_bytes(build)
    started = time.monotonic()
    text = xlsx.extract_xlsx(data)
    elapsed = time.monotonic() - started
    assert text is not None
    # Both values still land -- the cheap walk reads the same cells.
    assert "Alpha" in text, f"got: {text}"
    assert "Omega" in text, f"got: {text}"
    # 40 000 x 1 000 = 40 million coordinates; generous enough not to
    # flake on a loaded machine, tight enough that a bounding-rectangle
    # walk can't sneak back in unnoticed.
    assert elapsed < 3.0, f"sparse sheet took {elapsed:.3f}s -- the bounding-rectangle walk is back"


# ---------------------------------------------- decompression-bomb guards


def test_refuses_workbook_declaring_more_than_cap() -> None:
    # Decompression-bomb guard: a workbook whose entries declare more
    # decompressed bytes than the cap is refused before ever reaching
    # openpyxl.
    data = _fake_office_zip("xl/worksheets/sheet1.xml", "<x/>")
    assert xu.zip_declared_size_within(data, 1024) is True
    assert xu.zip_declared_size_within(data, 1) is False  # 4-byte entry > 1-byte cap
    assert xu.zip_declared_size_within(b"not a zip", 1024) is False


def test_refuses_workbook_inflating_more_than_cap() -> None:
    # The streaming guard counts ACTUAL inflated bytes, not headers.
    content = "x" * 100
    data = _fake_office_zip("xl/worksheets/sheet1.xml", content)
    assert xu.zip_inflated_size_within(data, 1024) is True
    assert xu.zip_inflated_size_within(data, 99) is False
    assert xu.zip_inflated_size_within(b"not a zip", 1024) is False


def test_refuses_workbook_with_lying_declared_sizes() -> None:
    # A crafted bomb declares a tiny uncompressed size but inflates far
    # past it; the declared-size fast path passes, so the streaming
    # re-count must be the one that refuses it.
    content = "x" * 100
    data = _fake_office_zip("xl/worksheets/sheet1.xml", content)
    patched = _patch_declared_size(data, 1)
    assert xu.zip_declared_size_within(patched, 64) is True, "fast path should pass"
    assert xu.zip_inflated_size_within(patched, 64) is False, "real guard must catch it"


def test_extract_xlsx_wired_to_the_decompression_guards(monkeypatch: pytest.MonkeyPatch) -> None:
    # Composition check beyond the three guard tests above: `extract_xlsx`
    # itself must consult the guards BEFORE ever handing bytes to
    # openpyxl, not just have them available. Patch the real module
    # constant down and confirm a workbook that easily clears the
    # production 512 MiB cap is refused once the cap is smaller than it.
    def build(ws):
        ws["A1"] = "Marketing"
        ws["B1"] = 12000

    data = _workbook_bytes(build)
    assert xlsx.extract_xlsx(data) is not None  # sanity: passes under the real cap

    monkeypatch.setattr(xlsx, "MAX_XLSX_DECOMPRESSED", 4)
    assert xlsx.extract_xlsx(data) is None


# ------------------------------------------------------------- formatting


def test_numeric_cell_display_has_no_trailing_dot_zero() -> None:
    # `str(float)` in Python always keeps a trailing `.0` on a whole
    # number (`str(100.0) == "100.0"`) -- umya's `cell.value()` shows a
    # user a plain "100", matching what they'd see in the cell. Verified
    # directly rather than assumed: a naive `str(value)` port would pass
    # the Rust-mirrored numeric test above (since `"12000.0"` still
    # contains `"12000"`) while silently emitting the wrong text.
    assert xlsx._format_cell_value(100.0) == "100"
    assert xlsx._format_cell_value(12000) == "12000"
    assert xlsx._format_cell_value(3.14) == "3.14"
    assert xlsx._format_cell_value(True) == "TRUE"
    assert xlsx._format_cell_value(False) == "FALSE"


def test_truncation_note_exact_message_format(monkeypatch: pytest.MonkeyPatch) -> None:
    # Port the exact message format: a row-cap truncation must read
    # `[sheet "{name}" truncated: read rows 1-{last_row} of {rows_total},
    # columns 1-{max_col} of {cols_total}]`.
    def build(ws):
        for row in range(1, 6):
            ws.cell(row=row, column=1, value=f"r{row}")

    data = _workbook_bytes(build)
    monkeypatch.setattr(xlsx, "MAX_ROWS", 2)
    text = xlsx.extract_xlsx(data)
    assert text is not None
    assert '[sheet "Sheet" truncated: read rows 1-2 of 5, columns 1-1 of 1]' in text, text
    assert "r3" not in text and "r4" not in text and "r5" not in text


def test_text_cap_stops_before_starting_the_next_row(monkeypatch: pytest.MonkeyPatch) -> None:
    # The text cap is checked at a row boundary after the sheet header has
    # already been appended. Keeping that order avoids a partial first row
    # while still reporting exactly what was not read.
    data = _workbook_bytes(lambda ws: ws.cell(row=1, column=1, value="hidden"))
    monkeypatch.setattr(xlsx, "MAX_TEXT_CHARS", 1)

    text = xlsx.extract_xlsx(data)

    assert text is not None
    assert "hidden" not in text
    assert '[sheet "Sheet" truncated: read rows 1-0 of 1, columns 1-1 of 1]' in text


# ------------------------------------------------------------------- misc


def test_non_zip_bytes_return_none() -> None:
    assert xlsx.extract_xlsx(b"not a zip at all") is None


def test_valid_zip_that_is_not_a_workbook_returns_none() -> None:
    # A structurally valid archive clears the size guards but must retain
    # the parser failure path rather than leaking openpyxl's exception.
    assert xlsx.extract_xlsx(_fake_office_zip("xl/worksheets/sheet1.xml", "<x/>")) is None


def test_zero_declared_extent_skips_the_cell_parser(monkeypatch: pytest.MonkeyPatch) -> None:
    # A zero extent is the pre-existing cheap skip path; it must not start
    # the private XML parser for a worksheet that advertises no cells.
    worksheet = SimpleNamespace(max_row=0, max_column=1, title="Ignored")
    output = xlsx._TextOutput()

    def parser_must_not_run(_: object):
        raise AssertionError("zero-extent sheet should not be parsed")

    monkeypatch.setattr(xlsx, "_iter_populated_cells", parser_must_not_run)
    xlsx._append_worksheet_text(worksheet, output)

    assert output.text() == ""


def test_stored_empty_cells_are_skipped_but_still_extend_the_true_bounds(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # The cell parser can expose empty stored cells and invalid coordinates
    # even though openpyxl's writer does not normally produce them. They do
    # not render a row, but their observed coordinates remain part of the
    # truthful truncation extent.
    cells = iter(
        [
            {"row": 0, "column": 1, "value": "invalid"},
            {"row": 3, "column": 4, "value": ""},
            {"row": 4, "column": 2, "value": "kept"},
        ]
    )
    monkeypatch.setattr(xlsx, "_iter_populated_cells", lambda _: cells)

    walk = xlsx._walk_populated_cells(object(), xlsx._TextOutput())

    assert walk.cells == ["", "kept"]
    assert (walk.max_row_seen, walk.max_col_seen) == (4, 4)


def test_lying_dimension_does_not_silently_drop_real_cells() -> None:
    # Regression: umya's `highest_row()`/`highest_column()` are computed
    # from the actual parsed cell collection -- there is no separate
    # "declared" bound in Rust for a workbook to lie about. openpyxl's
    # `<dimension>` element IS a separate, writer-supplied field, and a
    # non-openpyxl tool can understate it. A naive port that trusts
    # `ws.max_row`/`ws.max_column` for both the walk's inclusion bound and
    # the truncation report would silently drop every real cell beyond
    # the lied-about bound -- and, worse, wouldn't even flag it as
    # truncated, since the reported total is the same too-small lie.
    def build(ws):
        ws["A1"] = "InBounds"
        ws.cell(row=10, column=5, value="OutOfDeclaredBounds")

    data = _workbook_bytes(build)
    with zipfile.ZipFile(io.BytesIO(data)) as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")
    real_dim = sheet_xml[sheet_xml.index("<dimension ") : sheet_xml.index("/>", sheet_xml.index("<dimension ")) + 2]
    lied_xml = sheet_xml.replace(real_dim, '<dimension ref="A1:A1" />')
    assert lied_xml != sheet_xml, "dimension tag must actually change"

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = lied_xml.encode("utf-8")
            zout.writestr(item, content)
    lied_data = out.getvalue()

    text = xlsx.extract_xlsx(lied_data)
    assert text is not None
    assert "InBounds" in text, f"got: {text}"
    assert "OutOfDeclaredBounds" in text, f"cell beyond the lied dimension was dropped: {text}"


def test_blank_workbook_is_none_or_header_only() -> None:
    # An untouched openpyxl-written sheet still carries a `<dimension
    # ref="A1:A1"/>` (Excel's own convention for "no data yet"), so this
    # port emits a bare sheet header with no rows under it rather than
    # skipping the sheet outright -- see xlsx.py's `rows_total`/`cols_total`
    # comment. Either way there is no non-whitespace cell text.
    data = _workbook_bytes(lambda ws: None)
    text = xlsx.extract_xlsx(data)
    if text is not None:
        assert "Sheet" in text
        for line in text.splitlines():
            assert "|" not in line
