"""XLSX text extraction.

Port of `src-tauri/src/extraction/xlsx.rs`'s `extract_xlsx`. The zip-bomb
guards it also defines (`zip_declared_size_within`/`zip_inflated_size_within`)
are already ported onto the shared `arcelle_sidecar.docs.xml_utils` module
-- this file imports them rather than redefining them, exactly as the
migration brief asks.

------------------------------------------------------------------- why
The old xlsx reader only read `xl/sharedStrings.xml`, which interns STRING
cells only -- numeric cells live inline in each worksheet's own XML, so an
all-numeric sheet (a budget, a table of measurements) extracted to nothing
and the model saw the file as empty. `umya-spreadsheet` in Rust parses the
FULL workbook object model, so numbers/dates/formula results all land in
the extracted text; `openpyxl` is this port's equivalent, per the
migration plan's own recommendation.

---------------------------------------------------------------- bounds
Same three ceilings as the Rust source, deliberately wide (a real export
has to land WHOLE): `MAX_ROWS`/`MAX_COLS` bound the walked row/column
range; `MAX_TEXT_CHARS` bounds the accumulated output size -- compared
against Python string LENGTH (Unicode code points), not the Rust source's
UTF-8 BYTE length, since there is no cheap byte-length view of a growing
Python string buffer without re-encoding on every check. The 16 MiB budget
is generous enough on either side of that difference that it never changes
which sheets get truncated in any of this module's tests, and the
Python-side constant is itself named `MAX_TEXT_CHARS`, not bytes.
`MAX_XLSX_DECOMPRESSED` bounds how much the archive is allowed to
decompress to BEFORE the workbook is ever handed to openpyxl, which --
like umya -- fully loads into memory with no size guard of its own.

------------------------------------------------- walking populated cells ONLY
The Rust source's central perf fix is walking the sheet's POPULATED cells
(`ws.cells_sorted()`), never every coordinate in its rows x columns
bounding rectangle: a sheet holding two values whose highest cell sits at
row 40 000 / column 1 000 spans 40 million coordinates, and the original,
buggy version that walked all of them cost whole seconds with the room
lock held, over a few hundred bytes of real content.

The natural-looking Python translation -- `ws.iter_rows(values_only=True)`
in `read_only` mode -- does NOT have this property, confirmed empirically
(see `tests/test_docs_xlsx.py::test_sparse_sheet_costs_its_cells...`
which pins this down): read-only mode streams the XML row-by-row, but
`ReadOnlyWorksheet._cells_by_row`/`_get_row` still PAD every row they
yield out to the full column width. A row with no `<row>` element at all
reuses one cached all-`None` tuple rather than allocating a fresh one, but
the caller still receives that tuple and must scan every element of it to
find the populated ones -- so the walk's cost is proportional to
rows x columns, just with a cheap constant per empty cell rather than an
allocation. Measured directly against this file's own sparse-sheet
scenario (two real cells, one at row 40 000 / column 1 000): the plain
`iter_rows(values_only=True)` translation took ~1.0s -- comfortably under
the test's 3s budget, but nowhere near umya's ~20ms and not actually
cell-count-proportional; a genuinely sparse workbook nearer the real
`MAX_ROWS` x `MAX_COLS` ceiling (100 000 x 1 000) would be meaningfully
slower, reintroducing a shadow of the exact cost class this whole module
exists to avoid.

So this port reaches one layer deeper, into the same `WorkSheetParser`
that `ReadOnlyWorksheet._cells_by_row` itself is built on
(`openpyxl.worksheet._reader.WorkSheetParser`, fed the same private
`ws._get_source()` / `ws._shared_strings` / `wb.data_only` / `wb.epoch` /
`wb._date_formats` / `wb._timedelta_formats` that method already
assembles internally) to stream only the `<c>` elements actually present
in each `<row>` -- no per-row padding, no synthesized rows for gaps --
which is the real cell-count-proportional walk umya's `cells_sorted()`
gives Rust for free (measured at ~0.1ms for the same scenario). This
reaches past openpyxl's public surface into a private module; it is the
same private-attribute assembly `ReadOnlyWorksheet` already performs
internally, not a deeper coupling than that, and it is covered by this
file's own sparse-sheet timing test, so a future openpyxl upgrade that
changes these internals fails loudly here rather than silently
reintroducing the bounding-rectangle cost.

Containment note: like the Rust source, `extract_xlsx` itself only
guards the INITIAL workbook parse (`.ok()?` in Rust; a broad `except
Exception` here) and the zip-bomb precheck. A failure partway through the
cell walk (a corrupt XML fragment discovered mid-stream) is not
separately caught here, mirroring the Rust source's own structure: that
containment lives one layer up, in `extract_text`'s `contain_parser_panic`
wrapper, which is out of scope for this port (only `extract_xlsx` was
requested).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any, Iterator

import openpyxl
from openpyxl.worksheet._reader import WorkSheetParser

from arcelle_sidecar.docs.xml_utils import (
    zip_declared_size_within,
    zip_inflated_size_within,
)

# Bounds on how much of a workbook becomes searchable text. Deliberately
# wide -- a real export has to land WHOLE, and a tight cut would silently
# hide most of a big sheet from search and from the model while a viewer
# still showed every row. Whatever these do cut is ANNOUNCED in the text.
MAX_ROWS: int = 100_000
MAX_COLS: int = 1_000

# A second, absolute ceiling: 100 000 x 1 000 cells would be a multi-GB
# string. Whichever bound bites first, the truncation is reported.
MAX_TEXT_CHARS: int = 16 * 1024 * 1024

# openpyxl fully decompresses the workbook into an object model BEFORE the
# row/col bounds above apply, so a small zip bomb could balloon memory --
# exactly the risk umya has in Rust. Declared sizes can lie, so the
# declared-size check is only a fast path for honest oversized files; the
# streaming re-count of ACTUAL inflated bytes is the real guard. Generous
# on purpose -- a genuinely huge all-numeric sheet must still extract.
MAX_XLSX_DECOMPRESSED: int = 512 * 1024 * 1024

# NOTE: `rows_total`/`cols_total` inside `extract_xlsx` are NOT simply
# `ws.max_row`/`ws.max_column` (openpyxl's declared `<dimension>` metadata,
# which a non-openpyxl writer can understate) -- see the long comment at
# their computation site for why blindly trusting that field silently drops
# real trailing data, and how umya's `highest_row()`/`highest_column()`
# avoid the problem entirely by having no separate "declared" source at all.


@dataclass
class _TextOutput:
    """Accumulate extracted text while retaining the truncation budget."""

    parts: list[str] = field(default_factory=list)
    length: int = 0

    def append(self, text: str) -> None:
        self.parts.append(text)
        self.length += len(text)

    def text(self) -> str:
        return "".join(self.parts)


@dataclass
class _SheetWalk:
    """The observed extents and unfinished row from one worksheet walk."""

    cells: list[str] = field(default_factory=list)
    row_no: int = 0
    max_row_seen: int = 0
    max_col_seen: int = 0
    last_row: int | None = None


@dataclass(frozen=True)
class _SheetBounds:
    rows_total: int
    cols_total: int
    max_row: int
    max_col: int


def extract_xlsx(data: bytes) -> str | None:
    """Extract every populated cell of every worksheet as text, numbers and
    strings alike, in workbook sheet order. `None` for anything that isn't
    a readable workbook, over either decompression-bomb cap, or whose
    extracted text is all-whitespace.
    """
    if not _xlsx_is_safe_to_load(data):
        return None
    workbook = _load_xlsx(data)
    if workbook is None:
        return None
    text = _extract_workbook_text(workbook)
    return None if text.strip() == "" else text


def _xlsx_is_safe_to_load(data: bytes) -> bool:
    """Apply both archive-size checks in their original short-circuit order."""
    if not zip_declared_size_within(data, MAX_XLSX_DECOMPRESSED):
        return False
    return zip_inflated_size_within(data, MAX_XLSX_DECOMPRESSED)


def _load_xlsx(data: bytes) -> Any | None:
    """Load only readable workbooks; malformed archives are an empty result."""
    try:
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None


def _extract_workbook_text(workbook: Any) -> str:
    """Extract all sheets and close the reader even when a later walk fails."""
    output = _TextOutput()
    try:
        for worksheet in workbook.worksheets:
            _append_worksheet_text(worksheet, output)
    finally:
        workbook.close()
    return output.text()


def _append_worksheet_text(worksheet: Any, output: _TextOutput) -> None:
    """Append one worksheet, retaining its declared extent for the notice."""
    declared_rows, declared_cols = _declared_sheet_extent(worksheet)
    if _has_no_declared_extent(declared_rows, declared_cols):
        return
    output.append(f"[sheet: {worksheet.title}]\n")
    walk = _walk_populated_cells(worksheet, output)
    _append_sheet_end(worksheet.title, declared_rows, declared_cols, walk, output)


def _declared_sheet_extent(worksheet: Any) -> tuple[int, int]:
    """Read the optional writer-supplied extent without treating it as truth."""
    return worksheet.max_row or 0, worksheet.max_column or 0


def _has_no_declared_extent(rows: int, cols: int) -> bool:
    """Skip the cheap, known-empty worksheet case before parsing its XML."""
    return rows == 0 or cols == 0


def _walk_populated_cells(worksheet: Any, output: _TextOutput) -> _SheetWalk:
    """Stream only stored cells while tracking their true worksheet extent."""
    walk = _SheetWalk()
    for cell in _iter_populated_cells(worksheet):
        _record_cell_extent(walk, cell)
        if not _cell_is_within_limits(cell):
            continue
        if not _cell_has_value(cell):
            continue
        if not _append_cell_to_row(walk, cell, output):
            break
    return walk


def _record_cell_extent(walk: _SheetWalk, cell: dict) -> None:
    """Keep the real extent even for an ignored or out-of-budget cell."""
    walk.max_row_seen = max(walk.max_row_seen, cell["row"])
    walk.max_col_seen = max(walk.max_col_seen, cell["column"])


def _cell_is_within_limits(cell: dict) -> bool:
    """Accept only positive coordinates inside the fixed extraction bounds."""
    row = cell["row"]
    col = cell["column"]
    return 0 < row <= MAX_ROWS and 0 < col <= MAX_COLS


def _cell_has_value(cell: dict) -> bool:
    """Ignore stored-but-empty cells without adding an empty row."""
    value = cell["value"]
    return value is not None and value != ""


def _append_cell_to_row(walk: _SheetWalk, cell: dict, output: _TextOutput) -> bool:
    """Store a cell, flushing the preceding row on a row transition."""
    row = cell["row"]
    if row == walk.row_no:
        _store_cell(walk.cells, cell)
        return True
    return _start_row_with_cell(walk, cell, output)


def _start_row_with_cell(walk: _SheetWalk, cell: dict, output: _TextOutput) -> bool:
    """Open a new row before placing the cell that caused the transition."""
    if not _start_next_row(walk, cell["row"], output):
        return False
    _store_cell(walk.cells, cell)
    return True


def _start_next_row(walk: _SheetWalk, row: int, output: _TextOutput) -> bool:
    """Flush the preceding row and stop before starting an over-budget row."""
    _flush_pending_cells(walk, output)
    if output.length >= MAX_TEXT_CHARS:
        walk.last_row = walk.row_no
        return False
    walk.row_no = row
    return True


def _flush_pending_cells(walk: _SheetWalk, output: _TextOutput) -> None:
    """Render the open row with explicit placeholders for blank columns."""
    if not walk.cells:
        return
    output.append(" | ".join(walk.cells))
    output.append("\n")
    walk.cells.clear()


def _store_cell(cells: list[str], cell: dict) -> None:
    """Place one value at its one-based spreadsheet column."""
    col = cell["column"]
    missing_cells = col - len(cells)
    if missing_cells > 0:
        cells.extend([""] * missing_cells)
    cells[col - 1] = _format_cell_value(cell["value"])


def _append_sheet_end(
    title: str,
    declared_rows: int,
    declared_cols: int,
    walk: _SheetWalk,
    output: _TextOutput,
) -> None:
    """Finish a worksheet and visibly report any row or column truncation."""
    _flush_pending_cells(walk, output)
    bounds = _sheet_bounds(declared_rows, declared_cols, walk)
    last_row = _last_read_row(walk, bounds)
    if _sheet_was_truncated(last_row, bounds):
        output.append(
            f'[sheet "{title}" truncated: read rows 1-{last_row} of '
            f"{bounds.rows_total}, columns 1-{bounds.max_col} of {bounds.cols_total}]\n"
        )
    output.append("\n")


def _sheet_bounds(declared_rows: int, declared_cols: int, walk: _SheetWalk) -> _SheetBounds:
    """Combine declared and observed extents so understated metadata loses no cells."""
    rows_total = max(declared_rows, walk.max_row_seen)
    cols_total = max(declared_cols, walk.max_col_seen)
    return _SheetBounds(rows_total, cols_total, min(rows_total, MAX_ROWS), min(cols_total, MAX_COLS))


def _last_read_row(walk: _SheetWalk, bounds: _SheetBounds) -> int:
    """Use the cap-stop row when present, otherwise the complete row bound."""
    return bounds.max_row if walk.last_row is None else walk.last_row


def _sheet_was_truncated(last_row: int, bounds: _SheetBounds) -> bool:
    """Report either a row cap/text cap or a column cap."""
    return last_row < bounds.rows_total or bounds.cols_total > bounds.max_col


def _iter_populated_cells(ws: Any) -> Iterator[dict]:
    """Yield every actually-stored cell (`{"row", "column", "value", ...}`)
    of a read-only worksheet in its own XML order -- no filler for gaps
    within a row, no synthesized rows for entirely empty ones. See the
    module docstring for why this bypasses `ws.iter_rows()`.
    """
    with ws._get_source() as source:
        parser = WorkSheetParser(
            source,
            ws._shared_strings,
            data_only=ws.parent.data_only,
            epoch=ws.parent.epoch,
            date_formats=ws.parent._date_formats,
            timedelta_formats=ws.parent._timedelta_formats,
        )
        for _row_no, row_cells in parser.parse():
            yield from row_cells


def _format_cell_value(value: object) -> str:
    """Render a cell's raw Python value the way umya's `cell.value()`
    would -- numbers as plain numbers, with no artificial trailing `.0` on
    a value that is a whole number. Python's own `str(float)` always keeps
    that trailing zero (`str(100.0) == "100.0"`) -- worth calling out
    since it is exactly the kind of thing that looks right until checked:
    openpyxl's OWN writer never round-trips a whole-number float back as
    a literal `"100.0"` (it writes/reads `"100"`, landing as a Python
    `int`), so this branch only ever fires for a workbook written by some
    OTHER tool that stored an explicit trailing-`.0` numeric string --
    exactly the case this module's numeric test would miss if it only
    checked `"12000" in text`, which a stray `"12000.0"` also satisfies.
    """
    if isinstance(value, bool):
        # Excel shows a boolean cell as "TRUE"/"FALSE", not "1"/"0" --
        # checked first since `bool` is a subclass of `int` in Python.
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _format_float(value)
    return str(value)


def _format_float(value: float) -> str:
    """Avoid an artificial trailing `.0` on whole-number values."""
    if value.is_integer():
        return str(int(value))
    return str(value)
